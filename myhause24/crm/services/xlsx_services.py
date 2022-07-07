from django.db.models import Sum
from openpyxl import load_workbook
from openpyxl.styles import Side, Border, Font, Alignment
from openpyxl.writer.excel import save_virtual_workbook


def write_to_file(receipt, account, requisites, file, services, account_balance):

    data = {
        'total': services.aggregate(Sum('cost'))['cost__sum'],
        'totalDebt': float(str(account_balance.balance).replace('-', '')),
        'accountBalance': account_balance.balance,
        'invoiceNumber': receipt.number,
        'invoiceDate': receipt.date.strftime("%d.%m.%Y"),
        'invoiceMonth': f'{receipt.date_start.strftime("%d.%m")} - {receipt.date_end.strftime("%d.%m")}',
        'accountNumber': account.number,
        'invoiceAddress': f'{receipt.apartment.owner} '
                          f'{receipt.apartment.house.address}, '
                          f'квартира {receipt.apartment.number}',
        'payCompany': requisites,
    }

    template = load_workbook(filename=str(file.template.file))
    sheet = template.active

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in data.keys():
                sheet[cell.coordinate] = data[cell.value]

    start_service = 19

    thin = Side(border_style="thin", color="000000")

    for obj in services:
        sheet.insert_rows(start_service)

        sheet[f'A{start_service}'] = obj.services.title
        sheet[f'A{start_service}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet[f'A{start_service}'].font = Font(size=12, italic=False, color="000000")
        sheet.merge_cells(f'A{start_service}:B{start_service}')

        sheet[f'C{start_service}'] = obj.receipt.tariff.title
        sheet[f'C{start_service}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet[f'C{start_service}'].font = Font(size=12, italic=False, color="000000")
        sheet[f'C{start_service}'].alignment = Alignment(horizontal='right',
                                                         vertical='bottom')
        sheet.merge_cells(f'C{start_service}:D{start_service}')

        sheet[f'E{start_service}'] = obj.services.u_measurement.title
        sheet[f'E{start_service}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet[f'E{start_service}'].font = Font(size=12, italic=False, color="000000")
        sheet[f'E{start_service}'].alignment = Alignment(horizontal='right',
                                                         vertical='bottom')
        sheet.merge_cells(f'E{start_service}:F{start_service}')

        sheet[f'G{start_service}'] = obj.quantity
        sheet[f'G{start_service}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet[f'G{start_service}'].font = Font(size=12, italic=False, color="000000")
        sheet[f'G{start_service}'].alignment = Alignment(horizontal='right',
                                                         vertical='bottom')
        sheet.merge_cells(f'G{start_service}:H{start_service}')

        sheet[f'I{start_service}'] = obj.cost
        sheet[f'I{start_service}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet[f'I{start_service}'].font = Font(size=12, italic=False, color="000000")
        sheet[f'I{start_service}'].alignment = Alignment(horizontal='right',
                                                         vertical='bottom')
        sheet.merge_cells(f'I{start_service}:K{start_service}')

        start_service += 1

    sheet.merge_cells(f'A{start_service}:B{start_service}')
    sheet.merge_cells(f'C{start_service}:D{start_service}')
    sheet.merge_cells(f'E{start_service}:F{start_service}')
    sheet.merge_cells(f'G{start_service}:H{start_service}')
    sheet.merge_cells(f'I{start_service}:K{start_service}')

    return template

